#!/usr/bin/env python3
"""Create sample_input.xlsx with 15 tender questions."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

QUESTIONS = [
    (1, "Does your platform support TLS 1.2 or higher for all data in transit?"),
    (2, "Describe your disaster recovery plan and RPO/RTO targets."),
    (3, "What AI/ML capabilities does your platform offer?"),
    (4, "Is your organization SOC 2 Type II certified?"),
    (5, "Describe the overall system architecture and technology stack."),
    (6, "What is your pricing model for enterprise deployments?"),
    (7, "How do you handle data encryption at rest?"),
    (8, "What is your guaranteed uptime SLA?"),
    (9, "How do you ensure AI model fairness and prevent bias?"),
    (10, "Describe your GDPR compliance measures including data subject rights handling."),
    (11, "What integration options are available (APIs, webhooks, SSO)?"),
    (12, "Do you support multi-factor authentication for all user accounts?"),
    (13, "What cloud providers and regions do you support?"),
    (14, "Provide details on your data retention and disposal policies."),
    (15, "How does your platform handle horizontal scaling under peak load?"),
]

OUTPUT = Path(__file__).resolve().parent.parent / "data" / "sample_input.xlsx"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tender Questionnaire"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws["A1"] = "No"
    ws["B1"] = "Question"
    for cell in ["A1", "B1"]:
        c = ws[cell]
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin_border

    for row_idx, (num, question) in enumerate(QUESTIONS, 2):
        ws.cell(row=row_idx, column=1, value=num)
        ws.cell(row=row_idx, column=2, value=question)
        for col in [1, 2]:
            ws.cell(row=row_idx, column=col).border = thin_border
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 70

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
