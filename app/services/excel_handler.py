"""
Excel parsing and generation for tender questionnaires.

Infrastructure service — parses input Excel and generates output Excel.
"""

import logging
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def parse_tender_excel(file_content: bytes) -> list[dict]:
    """
    Parse an uploaded Excel file containing tender questions.

    Expected Excel format:
    - Column A (or "Question Number" / "No" / "#"): Question number
    - Column B (or "Question"): The tender question text

    The parser is flexible about column names (case-insensitive).
    If no question number column, auto-number sequentially.

    Returns:
        List of dicts with: question_number (int), original_question (str).

    Raises:
        ValueError: If file is empty or no questions found.
    """
    if not file_content:
        raise ValueError("Excel file is empty")

    try:
        df = pd.read_excel(BytesIO(file_content), engine="openpyxl", header=None)
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}") from e

    if df.empty:
        raise ValueError("Excel file is empty")

    question_col_idx: int | None = None
    number_col_idx: int | None = None
    data_start_row = 0

    question_aliases = ["question", "questions", "question text", "text", "q"]
    number_aliases = ["question number", "no", "no.", "#", "number", "num", "q#", "qno"]

    first_row = df.iloc[0]
    for idx in range(len(first_row)):
        val = first_row.iloc[idx]
        if pd.isna(val):
            continue
        s = str(val).strip().lower()
        for a in question_aliases:
            if a in s or s in a:
                question_col_idx = idx
                break
        for a in number_aliases:
            if a in s or s in a:
                number_col_idx = idx
                break

    if question_col_idx is not None or number_col_idx is not None:
        data_start_row = 1

    if question_col_idx is None:
        max_len, best_idx = 0.0, 0
        for idx in range(df.shape[1]):
            col = df.iloc[data_start_row:, idx].astype(str)
            avg_len = col.str.len().mean()
            if pd.notna(avg_len) and avg_len > max_len:
                max_len, best_idx = avg_len, idx
        question_col_idx = best_idx

    if number_col_idx is None:
        number_col_idx = 0 if question_col_idx != 0 else (1 if df.shape[1] > 1 else 0)

    questions: list[dict] = []
    for i in range(data_start_row, len(df)):
        row = df.iloc[i]
        q_text = row.iloc[question_col_idx] if question_col_idx < len(row) else None
        if pd.isna(q_text) or str(q_text).strip() == "":
            continue
        q_text = str(q_text).strip()
        if len(q_text) < 3:
            continue

        try:
            q_num_val = row.iloc[number_col_idx] if number_col_idx < len(row) else None
            if pd.isna(q_num_val):
                q_num = len(questions) + 1
            else:
                q_num = int(float(q_num_val))
        except (ValueError, TypeError):
            q_num = len(questions) + 1

        questions.append({"question_number": q_num, "original_question": q_text})

    if not questions:
        raise ValueError("No questions found in Excel file")

    logger.info("Parsed %d questions from Excel file", len(questions))
    return questions


def generate_output_excel(results: list[dict], summary: dict) -> bytes:
    """
    Generate an Excel file from processed tender results.

    Creates two sheets:
    1. "Responses" — One row per question with:
       Question Number | Original Question | Domain | Generated Answer | Confidence | Historical Match | Status

    2. "Summary" — Overall processing summary:
       Total Questions | Successful | Failed | Flagged | Overall Status

    Returns:
        bytes: Excel file content ready for streaming response.
    """
    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    headers = [
        "Question Number",
        "Original Question",
        "Domain",
        "Generated Answer",
        "Confidence",
        "Historical Match",
        "Status",
    ]
    ws = wb.active
    ws.title = "Responses"

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.get("question_number", ""))
        ws.cell(row=row_idx, column=2, value=r.get("original_question", ""))
        ws.cell(row=row_idx, column=3, value=r.get("domain", ""))
        ws.cell(row=row_idx, column=4, value=r.get("generated_answer", ""))
        conf = r.get("confidence", 0)
        conf_cell = ws.cell(row=row_idx, column=5, value=conf)
        if conf >= 0.8:
            conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif conf >= 0.5:
            conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.cell(row=row_idx, column=6, value="Yes" if r.get("has_historical_match") else "No")
        ws.cell(row=row_idx, column=7, value=r.get("status", ""))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["D"].width = 60

    ws_summary = wb.create_sheet("Summary", 0)
    summary_headers = ["Metric", "Value"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    summary_rows = [
        ("Total Questions", summary.get("total_questions", 0)),
        ("Successful", summary.get("successful", 0)),
        ("Failed", summary.get("failed", 0)),
        ("Flagged", summary.get("flagged", 0)),
        ("Overall Status", summary.get("overall_status", "")),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, 2):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
